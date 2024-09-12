import openpyxl
import pyautogui
import time
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
import threading
import keyboard
import requests
import os
import shutil
import subprocess
import json
from datetime import datetime, timedelta  # Adicione timedelta aqui
import sys
import re
import pdfplumber

# URL do ícone
icon_url = 'https://raw.githubusercontent.com/filipe131115/netnfnoah/main/icone.ico'

# Diretório temporário para salvar o ícone
icon_path = os.path.join(os.getcwd(), 'icone.ico')

# Baixa o ícone da URL e salva no diretório atual (uma vez só)
if not os.path.exists(icon_path):  # Verifica se o ícone já foi baixado
    response = requests.get(icon_url)
    with open(icon_path, 'wb') as file:
        file.write(response.content)

# URL do arquivo de usuários
USUARIOS_URL = "https://raw.githubusercontent.com/filipe131115/netnfnoah/main/usuarios.json"

# Variáveis globais
autenticado = False
executando = False
parar = False
link_txt = "https://raw.githubusercontent.com/filipe131115/netnfnoah/main/banco.txt"
arquivo_excel = ""
VERSAO_ATUAL = "1.07"
REPO_URL = "https://raw.githubusercontent.com/filipe131115/netnfnoah/main/version.txt"
EXECUTAVEL_URL = "https://bit.ly/AuttoPro"

def ler_usuarios():
    try:
        response = requests.get(USUARIOS_URL)
        response.raise_for_status()
        return response.json().get("usuarios", [])
    except requests.RequestException as e:
        messagebox.showerror("Erro", "Não foi possível carregar os dados dos usuários.")
        return []

def verificar_credenciais(usuario_input, senha_input):
    usuarios = ler_usuarios()
    for usuario in usuarios:
        if usuario['usuario'] == usuario_input and usuario['senha'] == senha_input:
            data_expiracao = datetime.strptime(usuario['data_expiracao'], "%Y-%m-%d")
            dias_restantes = (data_expiracao - datetime.now()).days
            if dias_restantes > 0:
                return True, dias_restantes
            else:
                return False, dias_restantes  # Conta vencida
    return False, None  # Usuário não encontrado

def centralizar_janela(janela, largura, altura):
    # Calcula a posição x e y para centralizar a janela
    tela_largura = janela.winfo_screenwidth()
    tela_altura = janela.winfo_screenheight()
    x = (tela_largura // 2) - (largura // 2)
    y = (tela_altura // 2) - (altura // 2)
    janela.geometry(f"{largura}x{altura}+{x}+{y}")
    
    
def mostrar_login():
    global autenticado
    login_window = tk.Tk()
    login_window.title("Login")
    centralizar_janela(login_window, 300, 230)  # Centraliza a janela de login
    login_window.configure(bg='#2e3b4e')
    login_window.iconbitmap(icon_path)
    
    # Define o ícone da janela de login a partir do arquivo baixado
    
    
    login_window.iconbitmap(icon_path)
    

    def verificar_login():
        usuario = entry_usuario.get()
        senha = entry_senha.get()
        valido, dias_restantes = verificar_credenciais(usuario, senha)

        if valido:
            data_expiracao = (datetime.now() + timedelta(days=dias_restantes)).strftime("%d/%m/%Y")
            messagebox.showinfo("Sucesso", f"Login bem-sucedido!\nData de expiração: {data_expiracao}")
            autenticado = True
            login_window.destroy()
            iniciar_janela_principal()
        else:
            if dias_restantes is not None:
                if dias_restantes <= 0:
                    messagebox.showerror("Erro", "Sua conta está vencida. Entre em contato com o suporte (71)99669-8662.")
                else:
                    messagebox.showerror("Erro", "Usuário ou senha incorretos.")
            else:
                messagebox.showerror("Erro", "Usuário ou senha incorretos.")

            if dias_restantes is not None:
                label_dias_restantes.config(text=f"{dias_restantes} dias restantes até a expiração." if dias_restantes > 0 else "Licença expirada!")

    def fechar_janela():
        login_window.destroy()
        sys.exit()

    login_window.protocol("WM_DELETE_WINDOW", fechar_janela)

    label_usuario = tk.Label(login_window, text="Usuário:", bg='#2e3b4e', fg='white')
    label_usuario.pack(pady=5)
    entry_usuario = tk.Entry(login_window, bg='white', fg='black')
    entry_usuario.pack(pady=5)

    label_senha = tk.Label(login_window, text="Senha:", bg='#2e3b4e', fg='white')
    label_senha.pack(pady=5)
    entry_senha = tk.Entry(login_window, show='*', bg='white', fg='black')
    entry_senha.pack(pady=5)

    botao_login = tk.Button(login_window, text="Login", command=verificar_login, bg='#4caf50', fg='white')
    botao_login.pack(pady=10)

    label_dias_restantes = tk.Label(login_window, bg='#2e3b4e', fg='white')
    label_dias_restantes.pack(pady=5)

    # Função para verificar login
    def verificar_login_event(event=None):
        verificar_login()

    # Vincula a tecla Enter ao campo de senha
    entry_senha.bind('<Return>', verificar_login_event)
    
    # Muda o foco para o campo de senha ao pressionar Enter no campo de usuário
    entry_usuario.bind('<Return>', lambda event: entry_senha.focus_set())

    login_window.mainloop()
    

def iniciar_janela_principal():
    if not autenticado:
        return

    principal_window = tk.Tk()
    principal_window.title("Janela Principal")
    centralizar_janela(principal_window, 600, 400)  # Centraliza a janela principal
    principal_window.configure(bg='#2e3b4e')

    # Adicione aqui os componentes da janela principal do seu aplicativo

    principal_window.mainloop()

# Chama a tela de login ao iniciar o aplicativo
mostrar_login()

# URL do arquivo .txt
link_txt = "https://raw.githubusercontent.com/filipe131115/netnfnoah/main/banco.txt"

# Função para atualizar o arquivo .txt
def atualizar_arquivo_txt():
    try:
        response = requests.get(link_txt)
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida
        with open("banco.txt", "wb") as f:
            f.write(response.content)
        print("Arquivo 'banco.txt' atualizado com sucesso.")
        messagebox.showinfo("Atualização", "Banco de dados atualizado com sucesso.")
    except Exception as e:
        print(f"Erro ao atualizar o arquivo: {e}")
        messagebox.showerror("Erro", f"Erro ao atualizar o banco de dados: {e}")


# Chama a função ao iniciar o programa
atualizar_arquivo_txt()

# Função para ler o arquivo .txt de um link externo e criar um dicionário de códigos e resultados
def ler_txt_link(link, progresso_callback):
    dados = {}
    try:
        response = requests.get(link)
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida
        linhas = response.text.splitlines()
        total_linhas = len(linhas)

        for idx, line in enumerate(linhas):
            partes = line.strip().split(':')  # Assumindo que o arquivo é no formato "código:resultado"
            if len(partes) == 2:
                codigo, resultado = partes
                dados[codigo.strip()] = resultado.strip()
            else:
                print(f"Formato inválido na linha: {line.strip()}")

            # Atualiza a barra de progresso a cada 5000 linhas
            if (idx + 1) % 5000 == 0:
                progresso_callback(idx + 1, total_linhas)

        # Garante que a barra de progresso chegue a 100% após o término
        progresso_callback(total_linhas, total_linhas)

    except requests.RequestException as e:
        atualizar_status(f"Ocorreu um erro ao acessar o arquivo .txt: {e}")
    except Exception as e:
        atualizar_status(f"Ocorreu um erro ao processar o arquivo .txt: {e}")
    return dados

# Função para atualizar a barra de progresso e o rótulo de porcentagem
def atualizar_progresso(progresso, total):
    porcentagem = (progresso / total) * 100
    progresso_bar['value'] = porcentagem
    porcentagem_label.config(text=f"{porcentagem:.2f}% Concluído")
    root.update_idletasks()

# Função para atualizar a mensagem de status
def atualizar_status(mensagem):
    status_label.config(text=mensagem)
    root.update_idletasks()

# Função para copiar e colar dados de A, B e C, parando quando encontrar uma célula vazia
def copiar_e_colar():
    global executando, parar, link_txt, arquivo_excel

    if not arquivo_excel:
        messagebox.showerror("Erro", "Arquivo Excel não selecionado.")
        return

    try:
        # Adicione aqui: Lê a linha inicial do campo de entrada
        linha_inicial = int(entry_linha_inicial.get())  # Lê a linha inicial do campo de entrada
        
        executando = True
        parar = False

        # Mensagem para que o usuário alterne manualmente para o programa destino
        messagebox.showinfo("Atenção", "Por favor, alterne para o programa de destino. O processo começará em 5 segundos.")
        time.sleep(5)  # Espera 5 segundos para dar tempo de mudar para o programa de destino

        # Carrega a planilha de origem
        atualizar_status("Carregando planilha Excel...")
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb.active

        # Lê os dados do link externo com barra de progresso
        atualizar_status("Lendo banco de dados...")
        progresso_bar['value'] = 0
        porcentagem_label.config(text="0.00% Concluído")
        dados_txt = ler_txt_link(link_txt, atualizar_progresso)
        if not dados_txt:
            messagebox.showerror("Erro", "Não foi possível ler o arquivo .txt.")
            wb.close()
            return
           
        progresso_bar['value'] = 100  # Define a barra de progresso como completa
        porcentagem_label.config(text="100.00% Concluído")

        for _ in range(1):  # Apenas uma vez, já que a quantidade de vezes foi removida
            if parar:
                atualizar_status("O processo foi interrompido.")
                messagebox.showinfo("Interrupção", "O processo foi interrompido com sucesso.")
                break

            for row in range(linha_inicial, sheet.max_row + 1):  # Começa da linha inicial
                if parar:
                    atualizar_status("O processo foi interrompido.")
                    messagebox.showinfo("Interrupção", "O processo foi interrompido com sucesso.")
                    break

                # Verifica se a célula da coluna A está vazia
                valor_a = sheet[f'A{row}'].value
                if valor_a is None:
                    atualizar_status("Célula vazia encontrada. Processo interrompido.")
                    break

                # Busca o resultado correspondente no arquivo .txt
                resultado = dados_txt.get(str(valor_a), '')
                if resultado:
                    # ======== Etapa 1: Digita o Resultado, Pressiona 1 Enter ========
                    pyautogui.write(resultado)  # Digita o resultado
                    pyautogui.press('enter')  # Pressiona Enter
                    pyautogui.press('enter')  # Pressiona Enter mais uma vez
                    time.sleep(1)

                    # ======== Etapa 2: Digita B, Pressiona 1 Enter ========
                    valor_b = sheet[f'B{row}'].value
                    if valor_b is not None:
                        pyautogui.write(str(valor_b))  # Digita o valor da célula B
                        pyautogui.press('enter')  # Pressiona Enter
                        time.sleep(1)

                    # ======== Etapa 3: Digita C, Pressiona 7 Enter ========
                    valor_c = sheet[f'C{row}'].value
                    if valor_c is not None:
                        valor_c_str = str(valor_c).replace('.', ',')  # Substitui o ponto por vírgula
                        pyautogui.write(valor_c_str)  # Digita o valor da célula C
                        for _ in range(7):  # Pressiona Enter 7 vezes
                            pyautogui.press('enter')
                        time.sleep(1)

        wb.close()
        if not parar:
            atualizar_status("Dados digitados com sucesso!")
        executando = False

    except Exception as e:
        atualizar_status(f"Ocorreu um erro: {e}")
        executando = False

# Função para iniciar a automação em uma thread separada
def iniciar_automacao():
    global executando
    if not executando:
        verificar_atualizacao()  # Verifica atualizações antes de iniciar
        threading.Thread(target=copiar_e_colar).start()

# Função para parar o processo
def parar_automacao():
    global parar
    parar = True
    atualizar_status("O processo foi interrompido.")
    messagebox.showinfo("Interrupção", "O processo foi interrompido com sucesso.")

# Função para selecionar o arquivo Excel
def selecionar_arquivo_excel():
    global arquivo_excel
    arquivo = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo:
        arquivo_excel = arquivo
        # Atualiza o texto do arquivo selecionado para uma mensagem simples
        label_arquivo_excel.config(text="Arquivo carregado com sucesso!")  # Mensagem de sucesso
        messagebox.showinfo("Sucesso", "Arquivo carregado com sucesso!")  # Mensagem de sucesso

        # Habilita o botão de iniciar se o link TXT estiver definido
        if link_txt:
            botao_iniciar.config(state=tk.NORMAL)

# Função para definir ou atualizar o link do arquivo TXT
def definir_link_txt():
    global link_txt
    novo_link = simpledialog.askstring("Link Banco", "Digite o link do banco:", initialvalue=link_txt)
    if novo_link:
        link_txt = novo_link
        atualizar_status("O link do banco foi atualizado com sucesso.")

# Função para verificar atualizações no aplicativo
def verificar_atualizacao():
    try:
        response = requests.get(REPO_URL)
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida
        nova_versao = response.text.strip()

        if nova_versao != VERSAO_ATUAL:
            atualizar_app()
        else:
            atualizar_status("O aplicativo está atualizado.")
    except requests.RequestException as e:
        atualizar_status(f"Erro ao verificar atualização: {e}")

# Função para atualizar o aplicativo
def atualizar_app():
    try:
        atualizar_status("Baixando nova versão...")
        response = requests.get(EXECUTAVEL_URL, stream=True)
        response.raise_for_status()

        # Salva o novo executável com o nome baseado na versão
        novo_executavel = f"Autto+{VERSAO_ATUAL}.exe"
        with open(novo_executavel, "wb") as f:
            shutil.copyfileobj(response.raw, f)

        atualizar_status("Atualização concluída!")
        
        # Informa ao usuário para fechar o aplicativo
        messagebox.showinfo("Atualização Completa", 
                            f"A atualização foi baixada com sucesso! "
                            f"Feche este aplicativo e inicie '{novo_executavel}'.")

        # O aplicativo atual pode ser fechado
        root.quit()  # Fecha a aplicação atual

    except Exception as e:
        atualizar_status(f"Erro ao atualizar o aplicativo: {e}")

# Caminho do executável antigo
executavel_antigo = "AuttoPro.exe"

# Verifica se o executável antigo existe e o remove
if os.path.exists(executavel_antigo):
    try:
        os.remove(executavel_antigo)
        print(f"Executável antigo '{executavel_antigo}' removido com sucesso.")
    except Exception as e:
        print(f"Erro ao remover o executável antigo: {e}")
        
        ## Função para renomear o executável e informar o usuário
def renomear_executavel():
    executavel_atual = os.path.basename(sys.argv[0])
    novo_nome = "AuttoPro.exe"
    
    if executavel_atual != novo_nome:
        try:
            os.rename(executavel_atual, novo_nome)
            messagebox.showinfo("Sucesso", f"Executável renomeado para '{novo_nome}'.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao renomear o executável: {e}")


# Chama a função de renomear ao iniciar o programa
renomear_executavel()


# Função para criar a calculadora
def criar_calculadora():
    calc = tk.Toplevel(root)
    calc.title("Calculadora")
    calc.geometry("600x600")
    calc.configure(bg='#2e3b4e')

    entrada = tk.Entry(calc, font=("Segoe UI", 18), bg='white', fg='black', justify='right')
    entrada.grid(row=0, column=0, columnspan=4, sticky="nsew")

    def adicionar_caractere(caractere):
        entrada.insert(tk.END, caractere)

    def calcular():
        try:
            # Substitui a vírgula por ponto
            entrada_valor = entrada.get().replace(',', '.')
            resultado = eval(entrada_valor)
            entrada.delete(0, tk.END)
            entrada.insert(tk.END, str(resultado))
        except Exception:
            entrada.delete(0, tk.END)
            entrada.insert(tk.END, "Erro")

    def limpar():
        entrada.delete(0, tk.END)

    def pressionar_enter(event):
        calcular()

    # Vincula a tecla Enter à função de calcular
    entrada.bind('<Return>', pressionar_enter)

    # Botões da calculadora
    botoes = [
        '7', '8', '9', '/',
        '4', '5', '6', '*',
        '1', '2', '3', '-',
        '0', 'C', '=', '+'
    ]

    linha = 1
    coluna = 0
    for botao in botoes:
        if botao == 'C':
            b = tk.Button(calc, text=botao, font=("Segoe UI", 18), command=limpar, bg='lightgray')
        elif botao == '=':
            b = tk.Button(calc, text=botao, font=("Segoe UI", 18), command=calcular, bg='lightgray')
        else:
            b = tk.Button(calc, text=botao, font=("Segoe UI", 18), command=lambda b=botao: adicionar_caractere(b), bg='lightgray')

        b.grid(row=linha, column=coluna, sticky="nsew", padx=1, pady=1)  # Espaçamento reduzido
        coluna += 1
        if coluna > 3:
            coluna = 0
            linha += 1

    for i in range(4):
        calc.grid_columnconfigure(i, weight=1)

    calc.grid_rowconfigure(0, weight=1)  # Para que a entrada ocupe espaço
    
   
# Função para mostrar a janela "Sobre"
def mostrar_sobre():
    texto_sobre = (
        "O AutoPro é um aplicativo desenvolvido para facilitar a automação da entrada de notas em processos "
        "com base em informações extraídas de planilhas Excel e bancos de dados. Ideal para quem precisa realizar "
        "tarefas repetitivas e deseja otimizar seu fluxo de trabalho.\n\n"
        "Com o AutoPro, você pode:\n\n"
        "- Selecionar arquivos Excel para ler dados a partir de planilhas.\n"
        "- Atualizar e definir links para bancos de dados.\n"
        "- Automatizar a entrada de dados em outros aplicativos, seguindo um processo específico para cada célula da planilha.\n"
        "- Monitorar o progresso da operação com uma barra de progresso detalhada e mensagens de status.\n\n"
        "Desenvolvido por Filipe Oliveira, o AutoPro visa simplificar e agilizar tarefas administrativas, proporcionando "
        "uma solução eficiente para usuários que necessitam de precisão e velocidade na entrada de notas.\n\n"
        "Ano de Desenvolvimento: 2024"
    )
    messagebox.showinfo("Sobre", texto_sobre)
    
    

# Função para mostrar a versão
def mostrar_versao():
    messagebox.showinfo("Versão", VERSAO_ATUAL)

# Função para monitorar os atalhos de teclado
def monitorar_atalhos():
    keyboard.add_hotkey('ctrl+alt+e', parar_automacao)  # Atalho para parar o processo
    keyboard.add_hotkey('ctrl+alt+r', iniciar_automacao)  # Atalho para iniciar o processo
    keyboard.add_hotkey('ctrl+alt+u', verificar_atualizacao)  # Atalho para verificar atualizações
    
def extrair_informacoes(pdf_path):
    dados = []
    ultima_data_vencimento = None
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    linhas = texto.split('\n')
                    
                    for linha in linhas:
                        if "data de vencimento:" in linha.lower():
                            match = re.search(r'\d{2}/\d{2}/\d{4}', linha)
                            if match:
                                ultima_data_vencimento = match.group(0)
                        
                        if "total:" in linha.lower() and ultima_data_vencimento:
                            match_total = re.search(r'[\d\.]+,\d{2}', linha)
                            if match_total:
                                total = match_total.group(0)
                                dados.append((ultima_data_vencimento, total))
    
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível ler o PDF: {e}")
    
    return dados

def gerar_relatorio_excel(dados):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Relatório"
    sheet.append(["Data de Vencimento", "Valor Total"])
    
    for data, total in dados:
        sheet.append([data, total])
    
    excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                filetypes=[("Excel Files", "*.xlsx")])
    if excel_path:
        wb.save(excel_path)
        messagebox.showinfo("Sucesso", "Relatório gerado com sucesso!")

def selecionar_arquivo_pdf():
    arquivo_pdf = filedialog.askopenfilename(title="Selecionar PDF", filetypes=[("Arquivos PDF", "*.pdf")])
    if arquivo_pdf:
        dados = extrair_informacoes(arquivo_pdf)
        if dados:
            gerar_relatorio_excel(dados)
        else:
            messagebox.showwarning("Aviso", "Não foi possível encontrar as informações no PDF.")

import os
import requests
import tkinter as tk

import os
import requests
import tkinter as tk

# Criação da interface gráfica
root = tk.Tk()
root.title("AutoPro")
root.geometry("700x590")
root.iconbitmap(icon_path)  # Reutiliza o ícone
root.configure(bg='#2e3b4e')  # Cor de fundo azul escuro

# Centraliza a janela principal
centralizar_janela(root, 700, 590)

# Estilo da barra de progresso
style = ttk.Style()
style.theme_use("clam")
style.configure("TProgressbar", troughcolor='lightgrey', background='deepskyblue', thickness=20)

# Frame para simular a barra de menu
menu_frame = tk.Frame(root, bg='#2e3b4e', bd=1, relief='raised')
menu_frame.pack(fill=tk.X)

# Labels do menu sem aspecto de botão
label_atualizar = tk.Label(menu_frame, text="Dados", bg='#2e3b4e', fg='white', cursor="hand2")
label_atualizar.pack(side=tk.LEFT, padx=5, pady=2)
label_atualizar.bind("<Button-1>", lambda e: definir_link_txt())

label_calc = tk.Label(menu_frame, text="Calculadora", bg='#2e3b4e', fg='white', cursor="hand2")
label_calc.pack(side=tk.LEFT, padx=5, pady=2)
label_calc.bind("<Button-1>", lambda e: criar_calculadora())

label_atualizar_ver = tk.Label(menu_frame, text="Atualizar", bg='#2e3b4e', fg='white', cursor="hand2")
label_atualizar_ver.pack(side=tk.LEFT, padx=5, pady=2)
label_atualizar_ver.bind("<Button-1>", lambda e: verificar_atualizacao())


label_sobre = tk.Label(menu_frame, text="Sobre", bg='#2e3b4e', fg='white', cursor="hand2")
label_sobre.pack(side=tk.LEFT, padx=5, pady=2)
label_sobre.bind("<Button-1>", lambda e: mostrar_sobre())


label_pdf = tk.Label(menu_frame, text="Pagamentos", bg='#2e3b4e', fg='white', cursor="hand2")
label_pdf.pack(side=tk.LEFT, padx=5, pady=2)
label_pdf.bind("<Button-1>", lambda e: selecionar_arquivo_pdf())


# Label de título
label_titulo = tk.Label(root, text="AutoPro: Automação de Pedidos", font=("Segoe UI", 16, 'bold'), bg='#2e3b4e', fg='white')
label_titulo.pack(pady=10)

# Label para exibir o caminho do arquivo Excel
label_arquivo_excel = tk.Label(root, text="", bg='#2e3b4e', fg='white')  # Inicialmente vazio
label_arquivo_excel.pack(pady=5)

# Botão para selecionar o arquivo Excel
botao_selecionar_arquivo = tk.Button(root, text="Selecionar Arquivo Excel", command=selecionar_arquivo_excel, bg='#4caf50', fg='white', borderwidth=2, relief='raised', width=20, height=2, font=("Segoe UI", 10))
botao_selecionar_arquivo.pack(pady=10)


# Rótulo e campo de entrada para a linha inicial
label_linha_inicial = tk.Label(root, text="Linha Inicial:", bg='#2e3b4e', fg='white')
label_linha_inicial.pack(pady=5)

# Declare a variável global
global entry_linha_inicial
entry_linha_inicial = tk.Entry(root, bg='white', fg='black')
entry_linha_inicial.pack(pady=5)
entry_linha_inicial.insert(0, "1")  # Definindo valor padrão como 1

# Barra de progresso
progresso_bar = ttk.Progressbar(root, length=400, mode='determinate')
progresso_bar.pack(pady=20)

# Label para mostrar a porcentagem de progresso
porcentagem_label = tk.Label(root, text="0.00% Concluído", bg='#2e3b4e', fg='white')
porcentagem_label.pack(pady=5)

# Frame para os botões de iniciar e parar
frame_botoes = tk.Frame(root, bg='#2e3b4e')
frame_botoes.pack(pady=20)

# Botão para iniciar a automação
botao_iniciar = tk.Button(frame_botoes, text="Iniciar Automação", command=iniciar_automacao, bg='#2196f3', fg='white', borderwidth=2, relief='raised', width=20, height=2, font=("Segoe UI", 10))
botao_iniciar.pack(side=tk.LEFT, padx=10)

# Botão para parar o processo
botao_parar = tk.Button(frame_botoes, text="Parar Automação", command=parar_automacao, bg='#f44336', fg='white', borderwidth=2, relief='raised', width=20, height=2, font=("Segoe UI", 10))
botao_parar.pack(side=tk.LEFT, padx=10)

# Label de status no rodapé
status_label = tk.Label(root, text="Pronto", bg='#2e3b4e', fg='white', anchor='w', padx=10, pady=5)
status_label.pack(side=tk.LEFT, fill=tk.X)



# Label para a versão no rodapé à direita
versao_label = tk.Label(root, text=f"Versão {VERSAO_ATUAL}", bg='#2e3b4e', fg='white', anchor='e', padx=10, pady=5)
versao_label.pack(side=tk.RIGHT)

# Inicia o monitoramento dos atalhos
monitorar_atalhos()
root.mainloop()